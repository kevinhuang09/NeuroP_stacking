# ======================================================================================================================
# 所有功能開關 (bool)，統一放在檔案最前面方便控制
disablePlotPopup = True  # 是否禁止彈出圖表視窗
useVscodeParentPath = True  # 使用pycharm, vscode進行編譯請開啟
b_saveFastaSeqCountStat = True  # 是否印出並儲存 fasta 序列數量統計
# ======================================================================================================================

import matplotlib

def setMatplotlibBackend(disablePopup):
    """disablePopup=True: 使用Agg backend，禁止彈出圖表視窗；disablePopup=False: 維持預設backend，正常跳出視窗"""
    if disablePopup:
        matplotlib.use('Agg')

setMatplotlibBackend(disablePlotPopup)

import sys, os

def setupParentPath(enableVscodeParentPath):
    """
    enableVscodeParentPath=True:
    使用Pycharm, Vscode進行編譯時，把上一層目錄加入sys.path，方便import套件
    """
    if enableVscodeParentPath:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        parent = os.path.join(current_dir, "..")
        sys.path.append(parent)

setupParentPath(useVscodeParentPath)

import copy
import csv
import json
from userPackage.Package_Encode import EncodeAllFeatures
from userPackage.FeatureStat import FeatureStat
from openpyxl import Workbook
from openpyxl.styles import Alignment

from sklearn.model_selection import train_test_split

from userPackage.LoadDataset import LoadDataset


from userPackage.FeatureConfig import (
    ifeatureDict,
    PfeatureDict,
    AMPfeatureDict,
    OVPfeatureDict,
    MotifBitVecfeatureDict,
    centerGDPDict,
    featureDict,
)

def buildOVPC_GAAC_formulaFeatureDict():
    """
    以 featureDict 為基礎，把原本各自獨立的 OVPC(ovpFeature)、GAAC(iFeature)、formula(ampFeature)
    三個開關關閉，取代成一個新群組 mergedFeature 底下的單一開關 OVPC_GAAC_formula 來代表這三者。
    其餘特徵開關維持不變。統計 feature type 數量時，三個獨立項目變成一個合併項目，理論上會少兩個。
    """
    newFeatureDict = copy.deepcopy(featureDict)

    newFeatureDict['iFeature']['GAAC'] = False
    newFeatureDict['ampFeature']['formula'] = False
    newFeatureDict['ovpFeature']['OVPC'] = False

    newFeatureDict['mergedFeature'] = {'OVPC,GAAC,formula': True}

    return newFeatureDict

# 這裡要用「featureDict 裡實際的 key」，大小寫要對上，不然 if feat in sub_dict 永遠找不到、關不掉
SINGLE_VALUE_FEATURE_NAME_LIST = [
    "length", "calculate_mw", "calculate_charge", "isoelectric_point",
    "instability_index", "aromaticity", "aliphatic_index", "hydrophobic",
    "aasi", "argos", "bulkiness", "charge_phys", "charge_acid",
    "flexibility", "gravy", "levitt_alpha", "mss", "polarity",
    "refractivity", "tm_tend", "boman_index", "eisenberg",
    "hopp_woods", "janin", "kytedoolittle", "SE", "charge_density"
]

def buildSingleValueCombineFeatureDict():
    """
    將單一數值型理化性質特徵關閉，
    改為在 mergedFeature 底下新增單一開關 SingleValueCombine。
    """
    newFeatureDict = copy.deepcopy(featureDict)

    print(f"singlevalue length : {len(SINGLE_VALUE_FEATURE_NAME_LIST)}")
    # 自動巡訪並關閉原始群組中的對應特徵開關
    for group, sub_dict in newFeatureDict.items():
        if isinstance(sub_dict, dict):
            for feat in SINGLE_VALUE_FEATURE_NAME_LIST:
                if feat in sub_dict:
                    if isinstance(sub_dict[feat], list):
                        sub_dict[feat][0] = False  # list 型參數只切換開關位，保留其餘參數
                    else:
                        sub_dict[feat] = False

    # 在 mergedFeature 加入合併開關 (避免覆蓋已存在的 mergedFeature)
    # if 'mergedFeature' not in newFeatureDict:
    #     newFeatureDict['mergedFeature'] = {}
    newFeatureDict['mergedFeature'][f'{",".join(map(str, SINGLE_VALUE_FEATURE_NAME_LIST))}'] = True

    return newFeatureDict

def countEnabledFeatureType(featureDict):
    """統計傳入的 featureDict 中，有多少個 feature type 被開啟(True)，並列出名稱"""
    enabledFeatureList = []

    for groupName, groupDict in featureDict.items():
        if 'Usage' in groupDict:
            if groupDict.get('Usage') is True:
                enabledFeatureList.append(f'{groupName}.Usage')
            continue
        for key, value in groupDict.items():
            if value is True or (isinstance(value, list) and value[0] is True):
                enabledFeatureList.append(f'{groupName}.{key}')

    return len(enabledFeatureList), enabledFeatureList
def buildAllOffFeatureDict(baseDict):
    """把 iFeature/pFeature/ampFeature/ovpFeature/centerGDPFeature 全部開關關閉，保留其餘參數結構"""
    offDict = copy.deepcopy(baseDict)
    for groupName in ['iFeature', 'pFeature', 'ampFeature', 'ovpFeature']:
        if groupName not in offDict:
            continue
        for key, value in offDict[groupName].items():
            if isinstance(value, list):
                value[0] = False
            else:
                offDict[groupName][key] = False
    offDict['centerGDPFeature']['Usage'] = False
    return offDict


def getRealEnabledFeatureTypeList(featureDict):
    """列出真正會產生欄位的 feature type（(groupName, key) tuple 清單）"""
    enabledList = []
    for groupName in ['iFeature', 'pFeature', 'ampFeature', 'ovpFeature']:
        if groupName not in featureDict:
            continue
        for key, value in featureDict[groupName].items():
            if value is True or (isinstance(value, list) and value[0] is True):
                enabledList.append((groupName, key))
    if featureDict['centerGDPFeature'].get('Usage') is True:
        enabledList.append(('centerGDPFeature', 'Usage'))
    return enabledList


def discoverFeatureTypeColumnMap(featureDict, sampleDataDict):
    """
    對每個真正開啟的 feature type，各自單獨開啟、encode 一次（只為了拿欄位名稱，不重算真正的特徵值），
    建立 feature type 名稱 -> 欄位名稱清單 的對照表，用來動態統計每個 feature type 實際產生的欄位數(feature size)。
    """
    realEnabledList = getRealEnabledFeatureTypeList(featureDict)
    allOffTemplate = buildAllOffFeatureDict(featureDict)

    featureTypeColumnMap = {}
    for groupName, key in realEnabledList:
        singleFeatureDict = copy.deepcopy(allOffTemplate)
        if isinstance(singleFeatureDict[groupName][key], list):
            singleFeatureDict[groupName][key][0] = True
        else:
            singleFeatureDict[groupName][key] = True

        singleEncodeObj = EncodeAllFeatures()
        singleEncodeObj.featureDict = singleFeatureDict
        singleDf = singleEncodeObj.dataEncodeOutPut(dataDict=sampleDataDict)
        columnList = [c for c in singleDf.columns if c != 'y']

        typeName = f'{groupName}.{key}'
        featureTypeColumnMap[typeName] = columnList

    return featureTypeColumnMap
def saveFeatureTypeReferenceTable(featureTypeColumnMap, savePath, removeList=None):
    """
    依 discoverFeatureTypeColumnMap 動態抓出的欄位清單，統計每個 feature type 實際的 feature size，
    輸出成 xlsx：欄位為 Type, Feature Type, Feature Size[, Feature Number After Filtering]。
    先依 Type 排序讓同樣的 type 集中在同一區塊（區塊內再依 feature size 由大到小排序），
    再把 Type 欄位中連續相同的值合併儲存格並置中。
    removeList 不為 None 時（feature filter 執行完之後），多輸出一欄 Feature Number After Filtering，
    統計每個 feature type 被過濾後還剩下多少欄位（未被 removeList 移除的欄位數）。
    """
    removeSet = set(removeList) if removeList is not None else None

    rows = []
    for typeName, columnList in featureTypeColumnMap.items():
        groupName, displayName = typeName.split('.', 1)
        row = [groupName, displayName, len(columnList)]
        if removeSet is not None:
            row.append(sum(1 for column in columnList if column not in removeSet))
        rows.append(row)

    # 依 Type 排序讓同樣的 type 集中在同一區，區內再依 feature size 由大到小排序
    rows.sort(key=lambda r: (r[0], -r[2]))

    wb = Workbook()
    ws = wb.active
    header = ['Type', 'Feature Type', 'Feature Size']
    if removeSet is not None:
        header.append('Feature Number After Filtering')
    ws.append(header)
    for row in rows:
        ws.append(row)

    # 把 Type 欄位中連續重複的值合併儲存格並置中
    typeCol = 1  # 'Type' 為第 1 欄
    headerRowOffset = 2  # 第 1 列為標題，資料從第 2 列開始
    centerAlign = Alignment(horizontal='center', vertical='center')

    startIdx = 0
    for i in range(1, len(rows) + 1):
        if i == len(rows) or rows[i][0] != rows[startIdx][0]:
            startRow = startIdx + headerRowOffset
            endRow = (i - 1) + headerRowOffset
            if endRow > startRow:
                ws.merge_cells(start_row=startRow, start_column=typeCol, end_row=endRow, end_column=typeCol)
            ws.cell(row=startRow, column=typeCol).alignment = centerAlign
            startIdx = i

    wb.save(savePath)
    print(f"Feature Type 對照表已儲存到 {savePath}")

def buildMergedFeatureTypeColumnMap(individualColumnMap):
    """
    以 discoverFeatureTypeColumnMap 針對『合併前』featureDict 算出的個別 feature type 欄位對照表為基礎，
    組出合併後兩個 merged feature type（OVPC_GAAC_formula 與 SingleValueCombine）各自的欄位清單，
    取代掉被合併掉的獨立項目，回傳完整的『合併後』欄位對照表。
    """
    mergedMap = copy.deepcopy(individualColumnMap)

    # 1. OVPC + GAAC + formula 合併成一項
    ovpcGaacFormulaColumns = []
    for typeName in ['ovpFeature.OVPC', 'iFeature.GAAC', 'ampFeature.formula']:
        ovpcGaacFormulaColumns.extend(mergedMap.pop(typeName, []))
    mergedMap['mergedFeature.OVPC,GAAC,formula'] = ovpcGaacFormulaColumns

    # 2. 27 個單一數值特徵合併成一項（這些 key 分散在 ampFeature 與 pFeature 兩個群組裡）
    singleValueColumns = []
    for feat in SINGLE_VALUE_FEATURE_NAME_LIST:
        for groupName in ['ampFeature', 'pFeature']:
            typeName = f'{groupName}.{feat}'
            if typeName in mergedMap:
                singleValueColumns.extend(mergedMap.pop(typeName))
                break

    mergedTypeName = f'mergedFeature.{",".join(SINGLE_VALUE_FEATURE_NAME_LIST)}'
    mergedMap[mergedTypeName] = singleValueColumns

    return mergedMap

# ======================================================================================================================
# 基本路徑
mlDataPath = "../data/mlData/"  # 內含 data 檔案 ex : train_F390.csv, boruta 檔案 ex :Boruta-featRank-RF.csv
paramPath = "../data/param/"  # 內含檔案: featureTypeDict.pkl, normalize.pkl
featureStatPath = '../data/featureStat/'
dataName = 'NeuroP_1'

normalizeMethodList = ['standard']  # normalization 目前先用standard 可自行改list

def saveFastaSeqCountStat(enable, fastaStatPath, statItems, statLabel='Fasta', fileMode='w', writeHeader=True):
    """
    印出並儲存目前序列數量統計。
    statItems: [(label, csvName, seqDict), ...]，label 用於印出訊息，csvName 為寫入 csv 的檔名/欄位名稱
    fileMode='w' 會覆寫並寫入表頭，'a' 則接續寫在同一份 csv 後面（不重寫表頭）。
    enable=False 時跳過印出與存檔。
    """
    if not enable:
        return

    statSummary = ", ".join(f"{label}: {len(seqDict)}" for label, _, seqDict in statItems)
    print(f"[{statLabel} 統計] {statSummary}")

    with open(fastaStatPath, fileMode, encoding='utf-8') as f:
        if writeHeader:
            f.write("file,count\n")
        for _, csvName, seqDict in statItems:
            f.write(f"{csvName},{len(seqDict)}\n")
    print(f"{statLabel} 數量統計已儲存到{fastaStatPath}中")

# DataSet載入區
# 載入Main Dataset
MainDatasetNegFastaPath = "../data/MainDatasetNeg.fasta"
MainDatasetPosFastaPath = "../data/MainDatasetPos.fasta"

# 載入DS_Indp
DS_IndpNegFastaPath = "../data/DS_IndpNeg.fasta"
DS_IndpPosFastaPath = "../data/DS_IndpPos.fasta"

ldObj = LoadDataset(minSeqLength=5)
MainDatasetNegSeqDict = ldObj.readFasta(MainDatasetNegFastaPath)
MainDatasetPosSeqDict = ldObj.readFasta(MainDatasetPosFastaPath)
DS_IndpNegSeqDict = ldObj.readFasta(DS_IndpNegFastaPath)
DS_IndpPosSeqDict = ldObj.readFasta(DS_IndpPosFastaPath)

# 印出並儲存當前序列數量統計
fastaStatPath = "../data/fasta_seq_count.csv"
saveFastaSeqCountStat(enable=b_saveFastaSeqCountStat,
                      fastaStatPath=fastaStatPath,
                      statItems=[
                          ('MainDatasetNeg', os.path.basename(MainDatasetNegFastaPath), MainDatasetNegSeqDict),
                          ('MainDatasetPos', os.path.basename(MainDatasetPosFastaPath), MainDatasetPosSeqDict),
                          ('DS_IndpNeg', os.path.basename(DS_IndpNegFastaPath), DS_IndpNegSeqDict),
                          ('DS_IndpPos', os.path.basename(DS_IndpPosFastaPath), DS_IndpPosSeqDict),
                      ])

# ======================================================================================================================
# 將 MainDataset 依 9:1 切分為 DS_Train / DS_Val（neg, pos 各自切分以維持類別比例）
splitTestSize = 0.1
splitRandomState = 42

def splitSeqDict(seqDict, test_size, random_state):
    keys = list(seqDict.keys())
    trainKeys, valKeys = train_test_split(keys, test_size=test_size, random_state=random_state)
    return {k: seqDict[k] for k in trainKeys}, {k: seqDict[k] for k in valKeys}

DS_TrainNegSeqDict, DS_ValNegSeqDict = splitSeqDict(MainDatasetNegSeqDict, splitTestSize, splitRandomState)
DS_TrainPosSeqDict, DS_ValPosSeqDict = splitSeqDict(MainDatasetPosSeqDict, splitTestSize, splitRandomState)

# ======================================================================================================================
# 只需要少量序列來辨識每個 feature type 產生的欄位名稱，不需要整個 DS_Train
columnDiscoverySampleSize = 5
columnDiscoverySampleDataDict = {
    0: dict(list(DS_TrainNegSeqDict.items())[:columnDiscoverySampleSize]),
    1: dict(list(DS_TrainPosSeqDict.items())[:columnDiscoverySampleSize]),
    -1: None
}

# 印出並儲存 DS_Train/DS_Val 數量統計（接續寫在同一份 fasta_seq_count.csv 中）
saveFastaSeqCountStat(enable=b_saveFastaSeqCountStat,
                      fastaStatPath=fastaStatPath,
                      statLabel='DS_Train/DS_Val',
                      fileMode='a',
                      writeHeader=False,
                      statItems=[
                          ('DS_TrainNeg', 'DS_TrainNeg', DS_TrainNegSeqDict),
                          ('DS_TrainPos', 'DS_TrainPos', DS_TrainPosSeqDict),
                          ('DS_ValNeg', 'DS_ValNeg', DS_ValNegSeqDict),
                          ('DS_ValPos', 'DS_ValPos', DS_ValPosSeqDict),
                      ])

DS_TrainDataDict = {0 : DS_TrainNegSeqDict, 1 : DS_TrainPosSeqDict, -1 : None}
DS_IndpDataDict = {0 : DS_IndpNegSeqDict, 1 : DS_IndpPosSeqDict, -1 : None}
DS_ValDataDict = {0 : DS_ValNegSeqDict, 1 : DS_ValPosSeqDict, -1 : None}

def saveDataDictToCsv(dataDict, csvPath):
    """把 dataDict（{-1/0/1: {序列名稱: 序列}}）攤平存成 csv，欄位為 name, sequence, label"""
    os.makedirs(os.path.dirname(csvPath), exist_ok=True)
    with open(csvPath, 'w', encoding='utf-8') as f:
        f.write("name,sequence,label\n")
        rowCount = 0
        for label, seqDict in dataDict.items():
            if seqDict is None:
                continue
            for name, sequence in seqDict.items():
                f.write(f"{name},{sequence},{label}\n")
                rowCount += 1
    print(f"{csvPath} 已儲存，共 {rowCount} 筆")

DS_TrainCsvPath = featureStatPath + 'DS_Train.csv'
DS_IndpCsvPath = featureStatPath + 'DS_Indp.csv'
DS_ValCsvPath = featureStatPath + 'DS_Val.csv'
saveDataDictToCsv(DS_TrainDataDict, DS_TrainCsvPath)
saveDataDictToCsv(DS_IndpDataDict, DS_IndpCsvPath)
saveDataDictToCsv(DS_ValDataDict, DS_ValCsvPath)

count, names = countEnabledFeatureType(featureDict)
print(f"featureDict 啟用的 feature type 數量: {count}")

# 保留合併前、所有開關都還是真實狀態的版本，實際 encode 一定要用這份：
# buildOVPC_GAAC_formulaFeatureDict / buildSingleValueCombineFeatureDict 只是把 OVPC/GAAC/formula
# 與 27 個單一數值特徵「重新標記」成一個 mergedFeature 開關，方便統計 feature type 數量與報表顯示；
# 但 Package_Encode.py 不認得 mergedFeature 這個 key，若拿合併後、個別開關已被關掉的版本去 encode，
# 這些欄位會整個消失（等於 merge 掉的 feature type 完全沒有資料可以訓練）。
originalFeatureDict = copy.deepcopy(featureDict)

os.makedirs(featureStatPath, exist_ok=True)
featureTypeColumnMap = discoverFeatureTypeColumnMap(originalFeatureDict, columnDiscoverySampleDataDict)
# featureTypeReferenceTablePath = featureStatPath + f'featureType_reference_table_{dataName}.csv'
# saveFeatureTypeReferenceTable(featureTypeColumnMap=featureTypeColumnMap,
                            #   savePath=featureTypeReferenceTablePath)

featureDict = buildOVPC_GAAC_formulaFeatureDict()
count2, names2 = countEnabledFeatureType(featureDict)
print(f"OVPC_GAAC_formula 啟用的 feature type 數量: {count2}")

featureDict = buildSingleValueCombineFeatureDict()
count3, names3 = countEnabledFeatureType(featureDict)
print(f"singleValueCombine 啟用的 feature type 數量: {count3}")

# 儲存這三步驟的 feature type 數量統計到 csv 中
featureTypeStatPath = featureStatPath + 'featureTypeStatistics.csv'
with open(featureTypeStatPath, 'w', encoding='utf-8') as f:
    f.write("featureType Process,count\n")
    f.write(f"origin featureDict,{count}\n")
    f.write(f"after merged OVPC GAAC formula,{count2}\n")
    f.write(f"after merged all single value,{count3}\n")
print(f"feature type 統計已儲存到{featureTypeStatPath}中")

# 把獨立版本的欄位對照表，組成合併後的版本，再存檔
os.makedirs(featureStatPath, exist_ok=True)
mergedFeatureTypeColumnMap = buildMergedFeatureTypeColumnMap(featureTypeColumnMap)
featureTypeReferenceTablePath = featureStatPath + f'featureType_reference_table_{dataName}.xlsx'
saveFeatureTypeReferenceTable(featureTypeColumnMap=mergedFeatureTypeColumnMap,
                              savePath=featureTypeReferenceTablePath)

# 橫向的 feature type 明細表：第一列為 Feature Type 名稱，第二列為 feature size，非 mergedFeature 排前面、mergedFeature 排最後
featureTypeTablePath = featureStatPath + f'featureType_featureName_table_{dataName}.csv'
allTypeKeys = list(mergedFeatureTypeColumnMap.keys())
otherTypeKeys = [k for k in allTypeKeys if not k.startswith('mergedFeature')]
mergedTypeKeys = [k for k in allTypeKeys if k.startswith('mergedFeature')]
orderedTypeKeys = otherTypeKeys + mergedTypeKeys

featureTypesRow = ['Feature Type']
featureSizesRow = ['feature size']
for typeName in orderedTypeKeys:
    columnList = mergedFeatureTypeColumnMap[typeName]
    displayName = typeName.split('.', 1)[-1]
    featureTypesRow.append(displayName)
    featureSizesRow.append(len(columnList))

with open(featureTypeTablePath, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(featureTypesRow)
    writer.writerow(featureSizesRow)
print(f"{len(mergedFeatureTypeColumnMap)} 類 feature type 明細表已儲存到 {featureTypeTablePath}")

# ======================================================================================================================
# 依 feature type 拆分表格用的工具函式（encode 後、normalize 後都會用到）
def getFeatureTypeFileName(typeName):
    """把 feature type 名稱轉成適合當檔名的字串（mergedFeature 底下的合併項目名稱太長，改用簡短名稱）"""
    groupName, displayName = typeName.split('.', 1)
    if groupName == 'mergedFeature':
        if displayName == 'OVPC,GAAC,formula':
            displayName = 'OVPC_GAAC_formula'
        elif displayName == ','.join(SINGLE_VALUE_FEATURE_NAME_LIST):
            displayName = 'SingleValueCombine'
    return f'{groupName}.{displayName}'

def saveEncodeDfSplitByFeatureType(encodeDf, featureTypeColumnMap, saveDir):
    """
    把一份表格依 featureTypeColumnMap 拆分，每個 feature type 各自存成一個 csv（含 y 欄位）。
    欄位若因過濾而不存在於 encodeDf 中會自動略過；一個 feature type 底下欄位全被過濾掉則不輸出該檔案。
    """
    os.makedirs(saveDir, exist_ok=True)
    savedCount = 0
    for typeName, columnList in featureTypeColumnMap.items():
        existColumnList = [c for c in columnList if c in encodeDf.columns]
        if not existColumnList:
            continue
        selectColumnList = existColumnList + (['y'] if 'y' in encodeDf.columns else [])
        typeCsvPath = os.path.join(saveDir, f'{getFeatureTypeFileName(typeName)}.csv')
        encodeDf[selectColumnList].to_csv(typeCsvPath)
        savedCount += 1
    print(f"{savedCount} 個 feature type 的 csv 已儲存到 {saveDir}")

# ======================================================================================================================
# Encode
encodeObj = EncodeAllFeatures()

encodeObj.dataEncodeSetup(saveFeatureDict=featureDict,  # normalization 前傳出來
                          saveJsonPath=paramPath + f'{dataName}_featureTypeDict.json',  # 把 featureDict 存至 json 檔
                          loadJsonPath=None,  # 讀取 featureDict 的 pkl 檔
                          b_loadJson=False)  # True: 讀取 featureDict 的 pkl 檔 (loadJsonPath), False: 把 featureDict 存至 pkl 檔 (saveJsonPath)

# json 存的是「合併顯示」版本（給 Main_MLStkLv1.py 判斷有哪些 real enabled feature type 用），
# 但實際 encode 要換回 originalFeatureDict，OVPC/GAAC/formula/27 個單一數值特徵才會真的產生欄位
encodeObj.featureDict = originalFeatureDict

encodeDS_TrainDf = encodeObj.dataEncodeOutPut(dataDict = DS_TrainDataDict)
encodeDS_IndpDf = encodeObj.dataEncodeOutPut(dataDict = DS_IndpDataDict)
encodeDS_ValDf = encodeObj.dataEncodeOutPut(dataDict = DS_ValDataDict)

# 儲存 encode 之後的大表格
os.makedirs(featureStatPath, exist_ok=True)
encodeDS_TrainCsvPath = featureStatPath + f'encode_{dataName}_DS_Train.csv'
encodeDS_IndpCsvPath = featureStatPath + f'encode_{dataName}_DS_Indp.csv'
encodeDS_ValCsvPath = featureStatPath + f'encode_{dataName}_DS_Val.csv'
encodeDS_TrainDf.to_csv(encodeDS_TrainCsvPath)
encodeDS_IndpDf.to_csv(encodeDS_IndpCsvPath)
encodeDS_ValDf.to_csv(encodeDS_ValCsvPath)
print(f"encode 後的資料已儲存到 {encodeDS_TrainCsvPath}、{encodeDS_IndpCsvPath} 與 {encodeDS_ValCsvPath}")

# 依 feature type 拆分 encode 後的大表格，每個 feature type 各自存成一個 csv，train/indp/val 各自一個資料夾
featureTypeSplitBaseDir = featureStatPath + f'featureType_csv_{dataName}/'
saveEncodeDfSplitByFeatureType(encodeDS_TrainDf, mergedFeatureTypeColumnMap, featureTypeSplitBaseDir + 'DS_Train/')
saveEncodeDfSplitByFeatureType(encodeDS_IndpDf, mergedFeatureTypeColumnMap, featureTypeSplitBaseDir + 'DS_Indp/')
saveEncodeDfSplitByFeatureType(encodeDS_ValDf, mergedFeatureTypeColumnMap, featureTypeSplitBaseDir + 'DS_Val/')

# ======================================================================================================================
# normalization：依 normalizeMethodList 逐一執行，分別存檔給 Main_MLStkLv1.py 讀取
for normalizeMethod in normalizeMethodList:
    nmlzScalerPath = paramPath + f'{dataName}_{normalizeMethod}Scaler.pkl'

    trainNmlzDf = encodeObj.dataNormalization(encodeTrainDf=encodeDS_TrainDf,
                                              encodeIndpDf=None,  # train scaler存起來，indp 另外做
                                              normalization=normalizeMethod,
                                              saveNmlzScalerPklPath=nmlzScalerPath,
                                              loadNmlzScalerPklPath=None,
                                              b_loadPkl=False)
    indpNmlzDf = encodeObj.dataNormalization(encodeTrainDf=None,
                                             encodeIndpDf=encodeDS_IndpDf,
                                             normalization=normalizeMethod,
                                             saveNmlzScalerPklPath=None,
                                             loadNmlzScalerPklPath=nmlzScalerPath,
                                             b_loadPkl=True)  # indp test set 永遠使用 training set 存好的 NmlzScaler.pkl 檔
    valNmlzDf = encodeObj.dataNormalization(encodeTrainDf=None,
                                            encodeIndpDf=encodeDS_ValDf,
                                            normalization=normalizeMethod,
                                            saveNmlzScalerPklPath=None,
                                            loadNmlzScalerPklPath=nmlzScalerPath,
                                            b_loadPkl=True)  # DS_Val 同樣永遠使用 training set 存好的 NmlzScaler.pkl 檔

    trainNmlzCsvPath = featureStatPath + f'train_{dataName}_{normalizeMethod}.csv'
    indpNmlzCsvPath = featureStatPath + f'indp_{dataName}_{normalizeMethod}.csv'
    valNmlzCsvPath = featureStatPath + f'val_{dataName}_{normalizeMethod}.csv'
    trainNmlzDf.to_csv(trainNmlzCsvPath)
    indpNmlzDf.to_csv(indpNmlzCsvPath)
    valNmlzDf.to_csv(valNmlzCsvPath)
    print(f"[{normalizeMethod}] normalize 後的資料已儲存到 {trainNmlzCsvPath}、{indpNmlzCsvPath} 與 {valNmlzCsvPath}")

    # 依 feature type 拆分 normalize 後的大表格，每個 feature type 各自存成一個 csv，train/indp/val 各自一個資料夾
    nmlzFeatureTypeSplitBaseDir = featureStatPath + f'featureType_csv_{dataName}_{normalizeMethod}/'
    saveEncodeDfSplitByFeatureType(trainNmlzDf, mergedFeatureTypeColumnMap, nmlzFeatureTypeSplitBaseDir + 'DS_Train/')
    saveEncodeDfSplitByFeatureType(indpNmlzDf, mergedFeatureTypeColumnMap, nmlzFeatureTypeSplitBaseDir + 'DS_Indp/')
    saveEncodeDfSplitByFeatureType(valNmlzDf, mergedFeatureTypeColumnMap, nmlzFeatureTypeSplitBaseDir + 'DS_Val/')

    # ==================================================================================================================
    # Feature Stat 分析：找出數值過度集中（top1percent 過高）的 feature 並過濾掉，MotifBitVec 系列 feature 受保護不被過濾
    featureStatObj = FeatureStat(dataDf=trainNmlzDf)
    featureStatObj.sdAnalysis(saveFigPath=featureStatPath + f"sd_analysis_{dataName}_{normalizeMethod}.jpg")
    featureAnalysisXlsxPath = featureStatPath + f"featureAnalysis_{dataName}_{normalizeMethod}.xlsx"
    featureStatObj.featureValuePct_analysis(saveFinalExcel=featureAnalysisXlsxPath)

    filteredTrainNmlzDf, removeList = featureStatObj.processData(xlsxPath=featureAnalysisXlsxPath,
                                                                  columnName='top1percent', number='+0.98',
                                                                  protectFeatSubstringList=['MotifBitVec'])
    os.makedirs(mlDataPath, exist_ok=True)
    featureStatObj.processDataLog(logPath=mlDataPath + f'{dataName}_{normalizeMethod}_')

    filteredIndpNmlzDf = indpNmlzDf.drop(columns=removeList)  # DS_Indp 用訓練集算出的 removeList 同步過濾，欄位才會跟訓練集一致
    filteredValNmlzDf = valNmlzDf.drop(columns=removeList)  # DS_Val 用訓練集算出的 removeList 同步過濾，欄位才會跟訓練集一致

    # 儲存過濾後的大表格
    filterTrainNmlzPath = featureStatPath + f'filtered_train_{dataName}_{normalizeMethod}.csv'
    filterIndpNmlzPath = featureStatPath + f'filtered_indp_{dataName}_{normalizeMethod}.csv'
    filterValNmlzPath = featureStatPath + f'filtered_val_{dataName}_{normalizeMethod}.csv'
    filteredTrainNmlzDf.to_csv(filterTrainNmlzPath)
    filteredIndpNmlzDf.to_csv(filterIndpNmlzPath)
    filteredValNmlzDf.to_csv(filterValNmlzPath)

    # 被過濾掉的 feature 名稱清單，以及依所屬 feature type 分組的清單
    removeFeatureListPath = featureStatPath + f'remove_feature_list_{dataName}_{normalizeMethod}.json'
    with open(removeFeatureListPath, 'w', encoding='utf-8') as f:
        json.dump(removeList, f, ensure_ascii=False, indent=2)

    columnToFeatureType = {column: typeName for typeName, columnList in mergedFeatureTypeColumnMap.items() for column in columnList}
    removeFeatureTypeDict = {}
    for column in removeList:
        typeName = columnToFeatureType.get(column, 'unknown')
        removeFeatureTypeDict.setdefault(typeName, []).append(column)
    removeFeatureTypeListPath = featureStatPath + f'remove_featureType_list_{dataName}_{normalizeMethod}.json'
    with open(removeFeatureTypeListPath, 'w', encoding='utf-8') as f:
        json.dump(removeFeatureTypeDict, f, ensure_ascii=False, indent=2)

    print(f"[{normalizeMethod}] Feature Stat 過濾完成，共過濾掉 {len(removeList)} 個 feature（分屬 {len(removeFeatureTypeDict)} 個 feature type），"
          f"剩餘 {filteredTrainNmlzDf.shape[1]} 欄，結果已儲存到 {filterTrainNmlzPath}、{filterIndpNmlzPath}、{filterValNmlzPath}、"
          f"{removeFeatureListPath} 與 {removeFeatureTypeListPath}")

    # feature filter 完成後，把過濾後剩餘欄位數輸出成獨立的每個 normalizeMethod 一份對照表（新增 Feature Number
    # After Filtering 欄位），檔名依 normalizeMethod 區分，避免 normalizeMethodList 有多個方法時互相覆蓋
    featureTypeFilterSummaryPath = featureStatPath + f'featureType_reference_table_{dataName}_{normalizeMethod}.xlsx'
    saveFeatureTypeReferenceTable(featureTypeColumnMap=mergedFeatureTypeColumnMap,
                                  savePath=featureTypeFilterSummaryPath,
                                  removeList=removeList)

    # 依 feature type 拆分過濾後的大表格，每個 feature type 各自存成一個 csv，train/indp/val 各自一個資料夾
    filteredFeatureTypeSplitBaseDir = featureStatPath + f'featureType_csv_{dataName}_{normalizeMethod}_filtered/'
    saveEncodeDfSplitByFeatureType(filteredTrainNmlzDf, mergedFeatureTypeColumnMap, filteredFeatureTypeSplitBaseDir + 'DS_Train/')
    saveEncodeDfSplitByFeatureType(filteredIndpNmlzDf, mergedFeatureTypeColumnMap, filteredFeatureTypeSplitBaseDir + 'DS_Indp/')
    saveEncodeDfSplitByFeatureType(filteredValNmlzDf, mergedFeatureTypeColumnMap, filteredFeatureTypeSplitBaseDir + 'DS_Val/')